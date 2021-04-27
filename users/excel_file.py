import openpyxl

wb = openpyxl.load_workbook('file.xlsx')
# print(type(wb))

number_of_sheets = wb.sheetnames # returns a list with the names of the sheet
# print(number_of_sheets)

# GET CURRENT ACTIVE SHEET
active_sheet = wb.active.title
# print(active_sheet)


# GET OBJECT OF A PARTICULAR SHEET
sh1 = wb['Sheet1']
# print(sh1)
# print(type(sh1))


# ACCESSING THE DATA - #OPTION 1
data = sh1['B1'].value
wb['Sheet1']['A3'].value
# print(data)


# ACCESSING THE DATA - #OPTION 2
# print(sh1.cell(3,2).value)


# ACCESSING THE DATA - #OPTION 3
# print(sh1.cell(row=3, column=2).value)

# GETTING NUMBER OF ROWS AND COLUMN
row = sh1.max_row
column = sh1.max_column
# print(row)
# print(column)

# for i in range(1, row + 1):
#     for j in range(1, column + 1):
#         print(sh1.cell(i, j).value)


for i in range(1, column + 1):
    # print(sh1.cell(row=1, column=i).value)

