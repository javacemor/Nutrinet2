from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from django.contrib.auth.models import User
from django.shortcuts import render
from rest_framework import status

import openpyxl
import os

# Rest Framework Imports
from rest_framework.decorators import api_view

# local imports
from .serializers import ProductSerializer, GroceryListSerializer, GroFiltersSerializer
from .models import Product, GroceryList, GroFilters


@api_view(['GET', 'POST'])
def products(request):
    if request.method == 'GET':
        products = Product.objects.all()
        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = ProductSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'POST', 'DELETE', 'PUT'])
def my_products(request, token):
    try:
        user = Token.objects.get(key=token).user
    except Token.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        products = GroceryList.objects.filter(user=user)
        serializer = GroceryListSerializer(products, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        product_id = request.data['product_id']
        product = Product.objects.get(id=product_id)

        grocery_list = GroceryList.objects.filter(user=user)
        
        for prod in grocery_list:
            if prod.product.id == product_id:
                prod.units += 1
                prod.save()
                return Response({}, status=status.HTTP_201_CREATED)
        
        request.data['user'] = user.pk
        request.data['product'] = product.pk
        request.data['image_url'] = product.image_url
        request.data['product_name'] = product.product_name
        serializer = GroceryListSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            print('no valid', serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'PUT':
        products = GroceryList.objects.filter(user=user)
        product_id = request.data['product_id']
        action = request.data['action']
        weeks = request.data['weeks']

        for product in products:
            if product.id == product_id:
                product = product

                if action == 'increment':
                    product.units += 1
                elif action == 'decrement':
                    if product.units < 1 or product.units == 1:
                        product.delete()
                        return Response(status=status.HTTP_204_NO_CONTENT)
                    else:
                        product.units -= 1
                elif action == 'updateWeeks':
                    product.weeks = weeks

                product.save()
        return Response({}, status=status.HTTP_201_CREATED)
        
    elif request.method == 'DELETE':
        products = GroceryList.objects.filter(user=user)
        product_id = request.data['product_id']
        for product in products:
            if product.id == product_id:
                product.delete()

        return Response(status=status.HTTP_204_NO_CONTENT)



@api_view(['GET', 'PUT', 'DELETE', 'POST'])
def my_product_filters(request, token):
    try:
        user = Token.objects.get(key=token).user
    except Token.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        pass
        # serializer = UserSerializer(user)
        # return Response(serializer.data)
    elif request.method == 'POST':
        request.data['user'] = user.pk
        p = GroFilters.objects.get(user=2)
        serializer = GroFiltersSerializer(p, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        else:
            print('NOT VALID')
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'DELETE':
        pass
        # user.delete()
        # return Response(status=status.HTTP_204_NO_CONTENT)


# @api_view(['GET', 'PUT', 'DELETE'])
# def product_detail(request, pk):
#     try:
#         user = User.objects.get(pk=pk)
#     except User.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)
    
#     if request.method == 'GET':
#         serializer = UserSerializer(user)
#         return Response(serializer.data)
#     elif request.method == 'PUT':
#         serializer = UserSerializer(user, data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#     elif request.method == 'DELETE':
#         user.delete()
#         return Response(status=status.HTTP_204_NO_CONTENT)









BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
file = os.path.join(BASE_DIR, "products/file.xlsx")
wb = openpyxl.load_workbook(file)

# wb = openpyxl.load_workbook('file.xlsx')
sh1 = wb['Sheet1']
row = sh1.max_row
column = sh1.max_column

def testing(request): 
    # for i in range(1, row + 1):
    #     for j in range(1, column + 1):
    #         print(sh1.cell(i, j).value)

    product_list = []

    for x in range(1, 509):
        product_dict = {}
        for i in range(1, column + 1):
            key = sh1.cell(row=1, column=i).value
            value = sh1.cell(row=x + 1, column=i).value
            product_dict[key] = value
        product_list.append(product_dict)
    for a in product_list:
        code = a['code']
        product_name = a['product_name']
        generic_name = a['generic_name']
        quantity = a['quantity']
        unit_quantity = a['unit quantity']
        packaging = a['packaging']
        brands = a['brands']
        categories = a['categories']
        origins = a['origins']
        labels = a['labels']
        stores = a['stores']
        countries = a['countries']
        ingredients_text = a['ingredients_text']
        allergens_tags = a['allergens_tags']
        traces = a['traces']
        additives_n = a['additives_n']
        additives_tags = a['additives_tags']
        ingredients_from_palm_oil_n = a['ingredients_from_palm_oil_n']
        nutriscore_grade = a['nutriscore_grade']
        nova_group = a['nova_group']
        pnns_groups_1 = a['pnns_groups_1']
        pnns_groups_2 = a['pnns_groups_2']
        main_category = a['main_category']
        image_url = a['image_url']
        image_small_url = a['image_small_url']
        image_ingredients_url = a['image_ingredients_url']
        image_nutrition_url = a['image_nutrition_url']
        energy_kcal_100g = a['energy-kcal_100g']
        fat_100g = a['fat_100g']
        saturated_fat_100g = a['saturated-fat_100g']
        monounsaturated_fat_100g = a['monounsaturated-fat_100g']
        polyunsaturated_fat_100g = a['polyunsaturated-fat_100g']
        omega_3_fat_100g = a['omega-3-fat_100g']
        omega_6_fat_100g = a['omega-6-fat_100g']
        trans_fat_100g = a['trans-fat_100g']
        carbohydrates_100g = a['carbohydrates_100g']
        sugars_100g = a['sugars_100g']
        fiber_100g = a['fiber_100g']
        proteins_100g = a['proteins_100g']
        salt_100g = a['salt_100g']
        sodium_100g = a['sodium_100g']
        alcohol_100g = a['alcohol_100g']
        vitamin_a_100g = a['vitamin-a_100g']
        vitamin_d_100g = a['vitamin-d_100g']
        vitamin_e_100g = a['vitamin-e_100g']
        vitamin_c_100g = a['vitamin-c_100g']
        vitamin_b1_100g = a['vitamin-b1_100g']
        vitamin_b2_100g = a['vitamin-b2_100g']
        vitamin_b6_100g = a['vitamin-b6_100g']
        vitamin_b9_100g = a['vitamin-b9_100g']
        vitamin_b12_100g = a['vitamin-b12_100g']
        biotin_100g = a['biotin_100g']
        pantothenic_acid_100g = a['pantothenic-acid_100g']
        potassium_100g = a['potassium_100g']
        calcium_100g = a['calcium_100g']
        phosphorus_100g = a['phosphorus_100g']
        iron_100g = a['iron_100g']
        magnesium_100g = a['magnesium_100g']
        zinc_100g = a['zinc_100g']
        copper_100g = a['copper_100g']
        manganese_100g = a['manganese_100g']
        selenium_100g = a['selenium_100g']
        iodine_100g = a['iodine_100g']
        caffeine_100g = a['caffeine_100g']
        fruits_vegetables_nuts_estimate_100g = a['fruits-vegetables-nuts-estimate_100g']
        cocoa_100g = a['cocoa_100g']
        carbon_footprint_from_meat_or_fish_100g = a['carbon-footprint-from-meat-or-fish_100g']
        price = a['Price']
        
        Product.objects.create(
            code= code,
            product_name=product_name,
            generic_name=generic_name,
            quantity=quantity,
            unit_quantity=unit_quantity,
            packaging=packaging,
            brands=brands,
            categories=categories,
            origins=origins,
            labels=labels,
            stores=stores,
            countries=countries,
            ingredients_text=ingredients_text,
            allergens_tags=allergens_tags,
            traces=traces,
            additives_n=additives_n,
            additives_tags=additives_tags,
            ingredients_from_palm_oil_n=ingredients_from_palm_oil_n,
            nutriscore_grade=nutriscore_grade,
            nova_group=nova_group,
            pnns_groups_1=pnns_groups_1,
            pnns_groups_2=pnns_groups_2,
            main_category=main_category,
            image_url=image_url,
            image_small_url=image_small_url,
            image_ingredients_url=image_ingredients_url,
            image_nutrition_url=image_nutrition_url,
            energy_kcal_100g=energy_kcal_100g,
            fat_100g=fat_100g,
            saturated_fat_100g=saturated_fat_100g,
            monounsaturated_fat_100g=monounsaturated_fat_100g,
            polyunsaturated_fat_100g=polyunsaturated_fat_100g,
            omega_3_fat_100g=omega_3_fat_100g,
            omega_6_fat_100g=omega_6_fat_100g,
            trans_fat_100g=trans_fat_100g,
            carbohydrates_100g=carbohydrates_100g,
            sugars_100g=sugars_100g,
            fiber_100g=fiber_100g,
            proteins_100g=proteins_100g,
            salt_100g=salt_100g,
            sodium_100g=sodium_100g,
            alcohol_100g=alcohol_100g,
            vitamin_a_100g=vitamin_a_100g,
            vitamin_d_100g=vitamin_d_100g,
            vitamin_e_100g=vitamin_e_100g,
            vitamin_c_100g=vitamin_c_100g,
            vitamin_b1_100g=vitamin_b1_100g,
            vitamin_b2_100g=vitamin_b2_100g,
            vitamin_b6_100g=vitamin_b6_100g,
            vitamin_b9_100g=vitamin_b9_100g,
            vitamin_b12_100g=vitamin_b12_100g,
            biotin_100g=biotin_100g,
            pantothenic_acid_100g=pantothenic_acid_100g,
            potassium_100g=potassium_100g,
            calcium_100g=calcium_100g,
            phosphorus_100g=phosphorus_100g,
            iron_100g=iron_100g,
            magnesium_100g=magnesium_100g,
            zinc_100g=zinc_100g,
            copper_100g=copper_100g,
            manganese_100g=manganese_100g,
            selenium_100g=selenium_100g,
            iodine_100g=iodine_100g,
            caffeine_100g=caffeine_100g,
            fruits_vegetables_nuts_estimate_100g=fruits_vegetables_nuts_estimate_100g,
            cocoa_100g=cocoa_100g,
            carbon_footprint_from_meat_or_fish_100g=carbon_footprint_from_meat_or_fish_100g,
            price=price,
            )
   
    return render(request, 'products/testing.html')

